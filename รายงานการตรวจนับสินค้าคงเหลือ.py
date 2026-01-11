import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from io import BytesIO
from datetime import datetime
from zoneinfo import ZoneInfo
from collections import OrderedDict

upload_file = st.session_state.get("excel_file_1")
wb = load_workbook(uploaded_file, data_only=True)
ws = wb.active
st.write(ำก)